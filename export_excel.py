"""
export_excel.py
---------------
Gera o Excel de exportacao no formato EXATO do arquivo "Modelo_1.xlsx":

    Aba 1 "BD"        -> dados de referencia gerais (listas fixas:
                          doencas, pragas, cultivares, regionais/municipios,
                          plantas invasoras, classes de produto, etc.)
    Aba 2 "Total_Pr"  -> tabulacao com as 129 colunas, no mesmo layout
                          de cabecalho do modelo, preenchida com os
                          dados reais cadastrados no site.

Nao ha formatacao (cor, negrito, largura, congelar paineis etc.) - apenas
o conteudo, exatamente como solicitado.

Uso no app.py (sem mudancas):

    from export_excel import gerar_excel, orm_para_dict

    @app.route("/exportar_excel")
    def exportar_excel():
        todos = FormularioSoja.query.order_by(FormularioSoja.id).all()
        registros = [orm_para_dict(r) for r in todos]
        filepath = os.path.join("/tmp", "MesoIDR_Export.xlsx")
        gerar_excel(registros, filepath)
        return send_file(filepath, as_attachment=True,
                         download_name="MesoIDR_Exportacao.xlsx",
                         mimetype="application/vnd.openxmlformats-"
                                  "officedocument.spreadsheetml.sheet")

IMPORTANTE: o arquivo "bd_dados.json" (dados fixos da aba BD) precisa estar
na MESMA PASTA deste arquivo export_excel.py.
"""

from __future__ import annotations

import json
import os
from typing import Any

from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Constantes de dominio (usadas apenas para reconhecer "alvo" texto -> praga/
# doenca ao montar o dicionario de cada registro)
# ---------------------------------------------------------------------------
REGIOES_IDR = [
    "Noroeste", "Norte", "Oeste", "Sudoeste",
    "Centro Sul", "Centro", "Metropolitana e Litoral",
]

PRAGAS = [
    "Lagarta da soja (Anticarsia gemmatalis)",
    "Lagarta das vagens (Spodoptera spp.)",
    "Lagarta falsa medideira (Chrysodeixis includens)",
    "Lagartas do grupo Heliothinae",
    "Percevejo barriga verde (Dichelops spp.)",
    "Percevejo marrom (Euschistus heros)",
    "Percevejo verde (Nezara viridula)",
    "Percevejo verde pequeno (Piezodorus guildinii)",
    "Broca dos ponteiros (Crocidosema aporema)",
    "Mosca Branca",
    "Outros insetos praga",
    "Tamandua da soja (Sternechus subsignatus)",
    "Tripes",
    "Vaquinhas (Diabrotica/ Cerotoma/ Colapsis)",
]

ACAROS = [
    "Acaro-rajado (Tetranychus urticae)",
    "Acaro-verde (Mononychellus planki)",
    "Acaro-branco (Polyphagotarsonemus latus)",
    "Acaros-vermelhos (Tetranychus spp.)",
    "Outros acaros",
]

DOENCAS_FUNGICAS = [
    "Antracnose (Colletotrichum truncatum)",
    "Cancro da haste (Diaporthe spp.)",
    "Ferrugem asiatica (Phakopsora pachyrhizi)",
    "Mancha alvo (Corynespora cassiicola)",
    "Mancha de cercospora (Cercospora kikuchii)",
    "Mancha olho-de-ra (Cercospora sojina)",
    "Mancha parda (Septoria glycines)",
    "Mela ou requeima (Rhizoctonia solani)",
    "Mofo branco (Sclerotinia sclerotiorum)",
    "Mildio (Peronospora manshurica)",
    "Oidio (Microsphaera diffusa)",
    "Outras Doencas Fungicas",
]
DOENCAS_BACT = [
    "Crestamento bacteriano (Pseudomonas savastanoi pv. glycinea)",
    "Fogo selvagem (Pseudomonas syringae pv. tabaci)",
    "Pustula bacteriana (Xanthomonas axonopodis pv. glycines)",
    "Mancha bacteriana marrom (Curtobacterium flaccumfaciens pv. flaccumfaciens)",
]
DOENCAS = DOENCAS_FUNGICAS + DOENCAS_BACT

N_PULV = 7

_LAGARTAS = [p for p in PRAGAS if "Lagarta" in p]
_PERCEVEJOS = [p for p in PRAGAS if "Percevejo" in p]
_PRAGAS_NOMEADAS = _LAGARTAS + _PERCEVEJOS
_OUTRAS_PRAGAS = [p for p in PRAGAS if p not in _PRAGAS_NOMEADAS]

_FERRUGEM = [d for d in DOENCAS_FUNGICAS if "Ferrugem" in d]
_MANCHA_ALVO = [d for d in DOENCAS_FUNGICAS if "Mancha alvo" in d]
_OIDIO = [d for d in DOENCAS_FUNGICAS if "Oidio" in d or "Oídio" in d]
_DOENCAS_NOMEADAS = _FERRUGEM + _MANCHA_ALVO + _OIDIO
_DEMAIS_FUNGICAS = [d for d in DOENCAS_FUNGICAS if d not in _DOENCAS_NOMEADAS]
_DOENCAS_MENOS_FERRUGEM = [d for d in DOENCAS if d not in _FERRUGEM]


def _norm(txt) -> str:
    """minusculas, sem acento, para comparacao tolerante de texto"""
    if not txt:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(txt))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def _match_lista(alvo, lista) -> bool:
    """True se o texto 'alvo' corresponde a algum item de 'lista' (comparacao
    tolerante: normaliza acentos/maiusculas e aceita substring)."""
    if not alvo:
        return False
    a = _norm(alvo)
    for item in lista:
        n = _norm(item)
        if a == n or n in a or a in n:
            return True
    return False


# ---------------------------------------------------------------------------
# Conversao ORM -> dict
# ---------------------------------------------------------------------------
def orm_para_dict(r) -> dict:
    """
    Converte FormularioSoja (com Pulverizacoes) para dict usado por gerar_excel().

    Modelo Pulverizacao:
      .tipo           -> 'dessecacao' | '1' | '2' ... '7'
      .data           -> string 'YYYY-MM-DD'
      .classe_produto -> ex. "Inseticida, Fungicida"
      .alvo           -> ex. "Lagarta da soja (Anticarsia gemmatalis), Ferrugem..."
    """
    pulvs: dict[str, Any] = {}
    for p in (r.pulverizacoes or []):
        pulvs[str(p.tipo).strip()] = p

    def _split(text: str) -> list[str]:
        if not text:
            return []
        return [x.strip() for x in text.replace("\n", ",").split(",") if x.strip()]

    def _dae(n: int):
        obj = pulvs.get(f"pos_{n}")
        if not obj or not obj.data or not r.data_emergencia:
            return None
        try:
            from datetime import datetime
            dp = datetime.strptime(obj.data[:10], "%Y-%m-%d").date()
            de = datetime.strptime(r.data_emergencia[:10], "%Y-%m-%d").date()
            return (dp - de).days
        except Exception:
            return None

    d: dict[str, Any] = {}

    # Identificacao
    d["N"]               = r.id
    d["Numero_Produtor"] = r.numero_produtor
    d["Meso_IDR"]        = r.meso_idr
    d["Regiao"]          = r.regiao
    d["Municipio"]       = r.municipio
    d["Area_Soja"]       = r.area_soja
    d["Cultivar"]        = r.cultivar
    d["Bt"]              = r.bt
    d["Produtividade"]   = r.produtividade_media
    d["Dt_Plantio"]      = r.data_plantio
    d["Adversidade"]     = r.qual_adversidade if r.houve_adversidade == "SIM" else None
    d["Sinistro"]        = r.houve_adversidade

    # Monitoramento
    d["Conhec_MID"]  = r.conhecimento_mid
    d["Utiliza_MID"] = r.utiliza_mid
    d["Conhec_MIP"]  = r.conhecimento_mip
    d["Utiliza_MIP"] = r.utiliza_mip

    # Plantas invasoras (ate 3 categorias hoje: dessecacao / pre / pos;
    # a 4a coluna do modelo fica em branco pois o site nao coleta esse dado)
    herbs = [
        ("Herbicida", r.herbicida_dessecacao_alvo, r.herbicida_dessecacao_aplicacoes),
        ("Herbicida", r.herbicida_pre_alvo,         r.herbicida_pre_aplicacoes),
        ("Herbicida", r.herbicida_pos_alvo,         r.herbicida_pos_aplicacoes),
        ("Herbicida", getattr(r, "herbicida_pos_ns_alvo", None), getattr(r, "herbicida_pos_ns_aplicacoes", None)),
    ]
    for i, (cl, alv, nap) in enumerate(herbs, start=1):
        d[f"Herb_Cl{i}"]  = cl if alv else None
        d[f"Herb_Alv{i}"] = alv
        d[f"Herb_Nap{i}"] = nap

    # Dessecacao (evento especifico, com data e ate 3 alvos)
    dess = pulvs.get("pre_plantio")
    d["Dess_Sim"] = "SIM" if dess else "NAO"
    d["Dess_Dt"]  = dess.data if dess else None
    d["Dess_Cl"]  = dess.classe_produto if dess else None
    dess_alvos = _split(dess.alvo) if dess else []
    for i in range(1, 4):
        d[f"Dess_Alv{i}"] = dess_alvos[i - 1] if i <= len(dess_alvos) else None

    # Pulverizacoes 1-7 (ate 5 classes/alvos por aplicacao)
    for n in range(1, N_PULV + 1):
        obj = pulvs.get(f"pos_{n}")
        alvos   = _split(obj.alvo)           if obj else []
        classes = _split(obj.classe_produto) if obj else []

        d[f"P{n}_DAE"]  = _dae(n)
        d[f"P{n}_Data"] = obj.data if obj else None

        for k in range(1, 6):
            d[f"P{n}_Cl{k}"]  = classes[k - 1] if k <= len(classes) else None
            d[f"P{n}_Alv{k}"] = alvos[k - 1]   if k <= len(alvos)   else None

    # Outras
    d["Tto_Semente"] = r.tratamento_sementes
    d["SAL_CB"]      = r.sal_mistura
    d["Ctrl_Biol"]   = r.controle_biologico

    # FBN / Inoculacao
    d["Inoc_Usa"]   = r.inoculacao_sementes
    d["Inoc_Forma"] = r.forma_inoculacao
    d["Coinoc"]     = r.coinoculacao
    d["CoMo_Usa"]   = r.co_mo
    d["CoMo_Forma"] = r.co_mo_aplicacao

    return d


# ---------------------------------------------------------------------------
# Aba "BD" - dados de referencia gerais (identico ao Modelo_1.xlsx)
# ---------------------------------------------------------------------------
def _carregar_bd_dados() -> list[list]:
    """Le o arquivo bd_dados.json (deve estar ao lado deste .py)."""
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bd_dados.json")
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def _build_BD(wb: Workbook):
    ws = wb.create_sheet("BD")
    linhas = _carregar_bd_dados()
    for ri, linha in enumerate(linhas, start=1):
        for ci, val in enumerate(linha, start=1):
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)


# ---------------------------------------------------------------------------
# Layout da aba "Total_Pr" (129 colunas), identico ao Modelo_1.xlsx
# Cada item: (chave_no_dict, texto_do_cabecalho, formato)
# formato: "num" -> numero, "txt"/None -> texto
# ---------------------------------------------------------------------------
def _pulv_group_cols(n: int, texto_dae: str):
    cols = [
        (f"P{n}_DAE",  texto_dae, "num"),
        (f"P{n}_Data", "Data",    "txt"),
    ]
    for k in range(1, 6):
        cols += [
            (f"P{n}_Cl{k}",  "Classe do Produto", "txt"),
            (f"P{n}_Alv{k}", "Alvo",              "txt"),
        ]
    return cols


_TEXTO_DAE = {
    1: "1ª Pulverização   (DAE)",
    2: "2ª Pulverização (DAE)",
    3: "3ª Pulverização   (DAE)",
    4: "4ª Pulverização  (DAE)",
    5: "5ª Pulverização  (DAE)",
    6: "6ª Pulverização  (DAE)",
    7: "7ª Pulverização (DAE)",
}

_TITULO_TOTAL_PR = (
    "PLANILHA TABULAÇÃO DADOS QUESTIONÁRIOS APLICAÇÃO DEFENSIVOS PARA "
    "CONTROLE PRAGAS E DOENÇAS_PR_SAFRA 19_20_V1"
)

# Colunas 1-4: sem "chave" de grupo (label fica direto na linha 3, sem
# subcabecalho na linha 4), exatamente como no modelo.
_COLS_INICIAIS = [
    (None,               None,     None),   # col 1 - sem cabecalho (numero sequencial)
    ("_TABELA",          "Tabela", "txt"),   # col 2
    ("Numero_Produtor",  "N° P",   "txt"),   # col 3
    (None,               "Ordem ", None),    # col 4 - sem subcabecalho (numero sequencial)
]

_ID_COLS = [
    ("Meso_IDR",      "Meso_IDR",                    "txt"),
    ("Regiao",        "Região",                      "txt"),
    ("Municipio",     "Município",                   "txt"),
    ("Area_Soja",     "Área com  Soja (ha)",          "num"),
    ("Cultivar",      "Cultivar",                     "txt"),
    ("Bt",            "Bt",                           "txt"),
    ("Produtividade", "Produtividade Média (sc/ha)",  "num"),
    ("Dt_Plantio",    "Data Plantio",                 "txt"),
    ("Adversidade",   "Adversidade",                  "txt"),
    ("Sinistro",      "Sinistro",                     "txt"),
]

_MID_COLS = [
    ("Conhec_MID",  "Conhec. MID",  "txt"),
    ("Utiliza_MID", "Utiliza MID",  "txt"),
    ("Conhec_MIP",  "Conhec. MIP",  "txt"),
    ("Utiliza_MIP", "Utiliza MIP",  "txt"),
]

# 4 blocos de (Classe do Produto / Alvo / N° Aplicações)
_HERB_COLS = []
for _i in range(1, 5):
    _HERB_COLS += [
        (f"Herb_Cl{_i}",  "Classe do Produto", "txt"),
        (f"Herb_Alv{_i}", "Alvo",              "txt"),
        (f"Herb_Nap{_i}", "N° Aplicações",     "num"),
    ]

_DESS_COLS = [
    ("Dess_Sim",  " Pulverização na Dessecação  ", "txt"),
    ("Dess_Dt",   "Data",                          "txt"),
    ("Dess_Cl",   "Classe do Produto",             "txt"),
    ("Dess_Alv1", "Alvo_1",                        "txt"),
    ("Dess_Alv2", "Alvo_2",                        "txt"),
    ("Dess_Alv3", "Alvo_3",                        "txt"),
]

_OUTRAS_COLS = [
    ("Tto_Semente", "Tratamento de Semente",           "txt"),
    ("SAL_CB",      "Utilização de SAL + Inseticida",  "txt"),
    ("Ctrl_Biol",   "Utilizou Controle Biológico",     "txt"),
]

_INOC_COLS = [
    ("Inoc_Usa",   "Utiliza Inoculação", "txt"),
    ("Inoc_Forma", "Forma Inoculação",   "txt"),
    ("Coinoc",     "Coinoculação",       "txt"),
    ("CoMo_Usa",   "Utiliza Co e Mo",    "txt"),
    ("CoMo_Forma", "Forma Co e Mo",      "txt"),
]

# grupos = (texto_do_grupo_ou_None, lista_de_colunas)
GRUPOS_TOTAL_PR = [
    (None,                                                     _COLS_INICIAIS),
    (None,                                                     _ID_COLS),
    ("CONHECIMENTO MONITORAMENTO",                             _MID_COLS),
    ("3_Informação Plantas Invasoras",                         _HERB_COLS),
    ("4.0_INFORMAÇÃO _PULVERIZAÇÃO DESSECAÇÃO",                _DESS_COLS),
    ("4.1_INFORMAÇÃO _PRIMEIRA PULVERIZAÇÃO APÓS EMERGÊNCIA",  _pulv_group_cols(1, _TEXTO_DAE[1])),
    ("4.2_INFORMAÇÃO _SEGUNDA PULVERIZAÇÃO APÓS EMERGÊNCIA",   _pulv_group_cols(2, _TEXTO_DAE[2])),
    ("4.3_INFORMAÇÃO _TERCEIRA PULVERIZAÇÃO APÓS EMERGÊNCIA",  _pulv_group_cols(3, _TEXTO_DAE[3])),
    ("4.4_INFORMAÇÃO _QUARTA PULVERIZAÇÃO APÓS EMERGÊNCIA",    _pulv_group_cols(4, _TEXTO_DAE[4])),
    ("4.5_INFORMAÇÃO _QUINTA PULVERIZAÇÃO APÓS EMERGÊNCIA",    _pulv_group_cols(5, _TEXTO_DAE[5])),
    ("4.6_INFORMAÇÃO _SEXTA PULVERIZAÇÃO APÓS EMERGÊNCIA",     _pulv_group_cols(6, _TEXTO_DAE[6])),
    ("4.7_INFORMAÇÃO _SÉTIMA PULVERIZAÇÃO APÓS EMERGÊNCIA",    _pulv_group_cols(7, _TEXTO_DAE[7])),
    ("5.OUTRAS INFORMAÇÕES",                                   _OUTRAS_COLS),
    ("6.INOCULAÇÃO",                                           _INOC_COLS),
]

ALL_COLS: list[tuple] = []
for _, cols in GRUPOS_TOTAL_PR:
    ALL_COLS.extend(cols)

# coluna 129 (DY) fica em branco, igual ao modelo
ALL_COLS.append((None, None, None))

_CI: dict[str, int] = {}
for _idx, (key, _label, _fmt) in enumerate(ALL_COLS, start=1):
    if key:
        _CI[key] = _idx


def _build_total_pr(wb: Workbook, registros: list[dict]):
    ws = wb.create_sheet("Total_Pr")
    nc = len(ALL_COLS)

    # Linha 1: titulo (somente na coluna D, igual ao modelo)
    ws.cell(row=1, column=4, value=_TITULO_TOTAL_PR)

    # Linha 3: cabecalhos de grupo (so na primeira coluna de cada grupo)
    # Linha 4: subcabecalhos (uma celula por coluna)
    col = 1
    for grupo_texto, cols in GRUPOS_TOTAL_PR:
        if grupo_texto is not None:
            ws.cell(row=3, column=col, value=grupo_texto)
        for key, label, _fmt in cols:
            if label is not None:
                # colunas iniciais (Tabela/N°P) tem o texto na propria linha 3
                if grupo_texto is None and cols is _COLS_INICIAIS:
                    ws.cell(row=3, column=col, value=label)
                else:
                    ws.cell(row=4, column=col, value=label)
            col += 1

    # Linha 5 fica em branco (igual ao modelo); dados comecam na linha 6
    linha_inicial = 6
    for i, reg in enumerate(registros):
        ri = linha_inicial + i
        ws.cell(row=ri, column=1, value=i + 1)      # col 1 - sequencial
        ws.cell(row=ri, column=2, value="TB1.")      # col 2 - Tabela
        ws.cell(row=ri, column=4, value=i + 1)       # col 4 - Ordem
        for key, _label, _fmt in ALL_COLS:
            if not key or key in ("_TABELA",):
                continue
            ci = _CI.get(key)
            if ci:
                ws.cell(row=ri, column=ci, value=reg.get(key))


# ---------------------------------------------------------------------------
# Aba "Médias_Geral" - indicadores agregados (calculados em Python direto a
# partir dos registros, no lugar de reproduzir as formulas originais que
# dependiam da aba auxiliar "Contagem_Pragas" de 534 colunas).
#
# LIMITACOES CONHECIDAS (o site hoje nao coleta esses dados, entao ficam
# zerados/ausentes ate que o formulario seja ampliado):
#   - Quebra "Folha Larga / Folha Estreita" dos herbicidas (o modelo original
#     usa uma classificacao escolhida manualmente no questionario, que nao
#     existe no banco atual).
#   - 4a categoria de "Plantas Invasoras" (o site so registra 3: dessecacao,
#     pre-emergente e pos-emergente).
# ---------------------------------------------------------------------------
REGIOES_COLUNAS = [None] + REGIOES_IDR  # None = "PARANA" (todo o estado)


def _filtra(registros: list[dict], regiao: str | None = None, bt: str | None = None):
    out = registros
    if regiao is not None:
        out = [r for r in out if _norm(r.get("Regiao")) == _norm(regiao)]
    if bt is not None:
        out = [r for r in out if _norm(r.get("Bt")) == _norm(bt)]
    return out


def _aplicacoes_alvo(reg: dict, lista_alvo: list[str]):
    """Lista de DAE (dias apos emergencia) de cada pulverizacao (1-7) que
    tenha atingido pelo menos um alvo de 'lista_alvo'."""
    daes = []
    for n in range(1, N_PULV + 1):
        alvos = [reg.get(f"P{n}_Alv{k}") for k in range(1, 6)]
        if any(_match_lista(a, lista_alvo) for a in alvos if a):
            dae = reg.get(f"P{n}_DAE")
            if dae is not None:
                daes.append(dae)
    return daes


def _media(lst):
    lst = [x for x in lst if x is not None]
    return (sum(lst) / len(lst)) if lst else None


def _bloco_alvo(registros: list[dict], lista_alvo: list[str]) -> dict:
    """Calcula os 10 indicadores padrao (COM/SEM aplicacao, %, n aplicacoes,
    DAE medio/primeira/menor/maior) para um alvo (praga, doenca, etc.)."""
    total = len(registros)
    com = sem = 0
    n_aplic_all, n_aplic_aplic = [], []
    dae_pool, dae_primeira = [], []
    for reg in registros:
        daes = _aplicacoes_alvo(reg, lista_alvo)
        n_aplic_all.append(len(daes))
        if daes:
            com += 1
            n_aplic_aplic.append(len(daes))
            dae_pool.extend(daes)
            dae_primeira.append(min(daes))
        else:
            sem += 1
    return {
        "com": com,
        "sem": sem,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
        "n_aplic_total": _media(n_aplic_all),
        "n_aplic_aplicantes": _media(n_aplic_aplic),
        "dae_medio": _media(dae_pool),
        "dae_primeira": _media(dae_primeira),
        "menor_dae_primeira": min(dae_primeira) if dae_primeira else None,
        "maior_dae_primeira": max(dae_primeira) if dae_primeira else None,
    }


_METRICAS_ALVO = [
    ("N° Questionários COM Aplicação {N}", "com"),
    ("N° Questionários SEM Aplicação {N}", "sem"),
    ("% Questionários COM Aplicação {N}", "pct_com"),
    ("% Questionários SEM Aplicação {N}", "pct_sem"),
    ("N° aplicações para {N} Total", "n_aplic_total"),
    ("N° aplicações para {N} Aplicantes", "n_aplic_aplicantes"),
    ("DAE_Médio aplicação {N}", "dae_medio"),
    ("DAE_Primeira aplicação {N}", "dae_primeira"),
    ("Menor DAE_Primeira Aplicação {N}", "menor_dae_primeira"),
    ("Maior DAE_Primeira Aplicação {N}", "maior_dae_primeira"),
]

# blocos de alvo, na mesma ordem do arquivo original
_BLOCOS_ALVO = [
    ("PRAGAS", PRAGAS),
    ("LAGARTAS", _LAGARTAS),
    ("ANTICARSIA GEMMATALIS", [p for p in PRAGAS if "Anticarsia" in p]),
    ("SPODOPTERA ssp.", [p for p in PRAGAS if "Spodoptera" in p]),
    ("CHRYSODEIXIS INCLUDENS", [p for p in PRAGAS if "Chrysodeixis" in p]),
    ("Grupo HELIOTHINAE", [p for p in PRAGAS if "Heliothinae" in p]),
    ("PERCEVEJOS", _PERCEVEJOS),
    ("DICHELOPS sp.", [p for p in PRAGAS if "Dichelops" in p]),
    ("EUSCHISTUS HEROS", [p for p in PRAGAS if "Euschistus" in p]),
    ("NEZARA VIRIDULA", [p for p in PRAGAS if "Nezara" in p]),
    ("PIEZODORUS GUILDINI", [p for p in PRAGAS if "Piezodorus" in p]),
    ("OUTRAS PRAGAS", _OUTRAS_PRAGAS),
    ("ÁCAROS", ACAROS),
    ("OUTRAS PRAGAS + Ácaros", _OUTRAS_PRAGAS + ACAROS),
    ("DOENÇAS", DOENCAS),
    ("FERRUGEM", _FERRUGEM),
    ("MANCHA ALVO", _MANCHA_ALVO),
    ("OÍDIO", _OIDIO),
    ("DEMAIS DOENÇAS FÚNGICAS", _DEMAIS_FUNGICAS),
    ("DOENÇAS BACTERIANAS", DOENCAS_BACT),
    ("TODAS DOENÇAS MENOS FERRUGEM", _DOENCAS_MENOS_FERRUGEM),
]


def _bloco_simples(registros, campo, valor_sim="SIM", valor_nao="NAO"):
    """Contagem COM/SEM/SEM_RESPOSTA para um campo de resposta SIM/NAO."""
    com = sem = sem_resposta = 0
    for reg in registros:
        v = _norm(reg.get(campo))
        if v == _norm(valor_sim):
            com += 1
        elif v == _norm(valor_nao):
            sem += 1
        else:
            sem_resposta += 1
    total = len(registros)
    return {
        "respondido": com + sem,
        "com": com,
        "sem": sem,
        "sem_resposta": sem_resposta,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
    }


def _build_medias_geral(wb: Workbook, registros: list[dict]):
    ws = wb.create_sheet("Médias_Geral")

    # ---- cabecalho ----
    ws.cell(row=1, column=1, value="Item")
    ws.cell(row=1, column=2, value="N° Questionários com respostas")
    col = 3
    col_map = []  # (col_index, regiao_ou_None, bt_ou_None)
    for regiao in REGIOES_COLUNAS:
        nome = regiao if regiao else "PARANÁ"
        ws.cell(row=1, column=col, value=nome)
        ws.cell(row=2, column=col, value="Total")
        col_map.append((col, regiao, None))
        ws.cell(row=2, column=col + 1, value="Cultivares_Bt")
        col_map.append((col + 1, regiao, "SIM"))
        ws.cell(row=2, column=col + 2, value="Cultivares_Não Bt")
        col_map.append((col + 2, regiao, "NAO"))
        col += 3

    linha = [3]  # contador de linha mutavel

    def escreve_linha(label, valores_por_regiao_bt):
        """valores_por_regiao_bt: dict {(regiao,bt): valor}"""
        r = linha[0]
        ws.cell(row=r, column=1, value=label)
        for c, regiao, bt in col_map:
            v = valores_por_regiao_bt.get((regiao, bt))
            if v is not None:
                ws.cell(row=r, column=c, value=v)
        # coluna B = mesmo valor da PARANÁ/Total
        vb = valores_por_regiao_bt.get((None, None))
        if vb is not None:
            ws.cell(row=r, column=2, value=vb)
        linha[0] += 1

    def pula_linha(n=1):
        linha[0] += n

    def para_cada_grupo(fn_calc):
        """fn_calc(regs_filtrados) -> dict de metricas; roda para todas as
        combinacoes de regiao/bt e devolve {(regiao,bt): dict}"""
        out = {}
        for c, regiao, bt in col_map:
            regs = _filtra(registros, regiao=regiao, bt=bt)
            out[(regiao, bt)] = fn_calc(regs)
        return out

    # ---------------- Bloco A: identificacao geral ----------------
    resultados_id = para_cada_grupo(lambda regs: {
        "n_aplicados": len(regs),
        "area_soja": sum(r.get("Area_Soja") or 0 for r in regs),
        "area_media": _media([r.get("Area_Soja") for r in regs]),
        "produtividade": _media([r.get("Produtividade") for r in regs]),
        "com_sinistro": sum(1 for r in regs if _norm(r.get("Sinistro")) == "sim"),
        "sem_sinistro": sum(1 for r in regs if _norm(r.get("Sinistro")) == "nao"),
        "area_com_sinistro": sum((r.get("Area_Soja") or 0) for r in regs if _norm(r.get("Sinistro")) == "sim"),
    })

    def _get(chave):
        return {k: v[chave] for k, v in resultados_id.items()}

    escreve_linha("N° Questionários Aplicados", _get("n_aplicados"))
    n_total_geral = resultados_id[(None, None)]["n_aplicados"]

    def _pct(vals_key, base=n_total_geral):
        out = {}
        for k, v in resultados_id.items():
            n = v["n_aplicados"]
            out[k] = (n / n_total_geral) if n_total_geral else None
        return out

    escreve_linha("Percentual Questionários", _pct("n_aplicados"))
    escreve_linha("N° Questionários COM Relato de SINISTRO", _get("com_sinistro"))
    escreve_linha("% Questionários COM Relato de SINISTRO",
                  {k: (v["com_sinistro"] / v["n_aplicados"]) if v["n_aplicados"] else None
                   for k, v in resultados_id.items()})
    escreve_linha("N° Questionários SEM Relato de SINISTRO", _get("sem_sinistro"))
    escreve_linha("% Questionários SEM Relato de SINISTRO",
                  {k: (v["sem_sinistro"] / v["n_aplicados"]) if v["n_aplicados"] else None
                   for k, v in resultados_id.items()})
    escreve_linha("Área_SOJA (ha)", _get("area_soja"))
    escreve_linha("Área média cultivada (ha)", _get("area_media"))
    escreve_linha("Área_Soja COM Relato SINISTRO", _get("area_com_sinistro"))
    escreve_linha("% área_Soja COM Relato SINISTRO",
                  {k: (v["area_com_sinistro"] / v["area_soja"]) if v["area_soja"] else None
                   for k, v in resultados_id.items()})
    escreve_linha("Produtividade (sc/ha)", _get("produtividade"))
    pula_linha()

    # ---------------- Blocos B/C/D: pragas, acaros, doencas ----------------
    for nome, lista in _BLOCOS_ALVO:
        resultados = para_cada_grupo(lambda regs, lst=lista: _bloco_alvo(regs, lst))
        for template, chave in _METRICAS_ALVO:
            escreve_linha(template.format(N=nome), {k: v[chave] for k, v in resultados.items()})
        pula_linha()

    # ---------------- Bloco E: totais gerais de aplicacao ----------------
    res_pragas = para_cada_grupo(lambda regs: _bloco_alvo(regs, PRAGAS))
    res_doencas = para_cada_grupo(lambda regs: _bloco_alvo(regs, DOENCAS))
    escreve_linha("Número Total Aplicação Inseticida",
                  {k: v["n_aplic_total"] for k, v in res_pragas.items()})
    escreve_linha("Número Total Aplicação Total Fungicida",
                  {k: v["n_aplic_total"] for k, v in res_doencas.items()})
    escreve_linha("Número Total Aplicação Total (Inseticida + Fungicida)",
                  {k: (res_pragas[k]["n_aplic_total"] or 0) + (res_doencas[k]["n_aplic_total"] or 0)
                   for k in res_pragas})
    pula_linha()

    # ---------------- Bloco F: tratamento semente / SAL / ctrl biologico ----------------
    def _bloco_campo(campo, valor_sim="SIM", valor_nao="NAO"):
        return para_cada_grupo(lambda regs: _bloco_simples(regs, campo, valor_sim, valor_nao))

    for campo, titulo in [
        ("Tto_Semente", "TRATAMENTO SEMENTE"),
        ("SAL_CB", "Utilização de SAL para controle percevejos"),
        ("Ctrl_Biol", "CONTROLE BIOLÓGICO"),
    ]:
        r = _bloco_campo(campo)
        escreve_linha(f"N° Questionários {titulo}", {k: v["respondido"] for k, v in r.items()})
        escreve_linha(f"N° Questionários COM Utilização de {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM Utilização {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM RESPOSTA {titulo}", {k: v["sem_resposta"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # Inseticida na dessecação (classe do produto contem "Inseticida")
    r_ins_dess = para_cada_grupo(lambda regs: _bloco_simples_texto(regs, "Dess_Cl", "Inseticida"))
    escreve_linha("N° Questionários COM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["com"] for k, v in r_ins_dess.items()})
    escreve_linha("N° Questionários SEM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["sem"] for k, v in r_ins_dess.items()})
    escreve_linha("Percentual Questionários COM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["pct_com"] for k, v in r_ins_dess.items()})
    escreve_linha("Percentual de Questionários SEM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["pct_sem"] for k, v in r_ins_dess.items()})
    pula_linha()

    # ---------------- Bloco G: conhecimento/uso MID e MIP ----------------
    for campo, titulo in [
        ("Conhec_MID", "CONHECIMENTO MID"),
        ("Utiliza_MID", "USO MID"),
        ("Conhec_MIP", "CONHECIMENTO MIP"),
        ("Utiliza_MIP", "USO MIP"),
    ]:
        r = _bloco_campo(campo)
        escreve_linha(f"N° Questionários {titulo}", {k: v["respondido"] for k, v in r.items()})
        escreve_linha(f"N° Questionários COM {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM RESPOSTA {titulo}", {k: v["sem_resposta"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # ---------------- Bloco H: herbicidas (COM/SEM aplicacao, geral) ----------------
    # OBS: a quebra "Folha Larga/Folha Estreita" do arquivo original depende de
    # uma classificacao manual que o site ainda nao coleta - por isso essas
    # sub-linhas nao aparecem aqui (ver observacao no topo do arquivo).
    for campo_alvo, campo_nap, titulo in [
        ("Herb_Alv1", "Herb_Nap1", "Herbicida não seletivo na dessecação"),
        ("Herb_Alv2", "Herb_Nap2", "Herbicida Pré emergente"),
        ("Herb_Alv3", "Herb_Nap3", "Herbicida Pós emergente"),
    ]:
        r = para_cada_grupo(lambda regs, c=campo_alvo: _bloco_simples_preenchido(regs, c))
        escreve_linha(f"N° Questionários COM aplicação {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM aplicação {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM Aplicação {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM Aplicação {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # ---------------- Bloco I: inoculacao / coinoculacao / Co e Mo ----------------
    for campo, titulo in [
        ("Inoc_Usa", "USO INOCULAÇÃO"),
        ("Coinoc", "COINOCULAÇÃO"),
        ("CoMo_Usa", "USO Co Mo"),
    ]:
        r = _bloco_campo(campo)
        escreve_linha(f"N° Questionários RESPOSTA {titulo}", {k: v["respondido"] for k, v in r.items()})
        escreve_linha(f"N° Questionários COM {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM RESPOSTA {titulo}", {k: v["sem_resposta"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # forma de inoculacao (categorias de texto livre)
    r_forma = para_cada_grupo(lambda regs: {
        cat: sum(1 for r in regs if cat_lower in _norm(r.get("Inoc_Forma")))
        for cat, cat_lower in [
            ("industrial", "industrial"),
            ("caixa", "caixa"),
            ("misturador", "misturador"),
            ("betoneira", "betoneira"),
            ("lona", "lona"),
            ("sulco", "sulco"),
        ]
    })
    for cat, titulo in [
        ("industrial", "INOCULAÇÃO INDUSTRIAL"),
        ("caixa", "INOCULAÇÃO CAIXA PLANTADEIRA"),
        ("misturador", "INOCULAÇÃO MISTURADOR SEMENTE"),
        ("betoneira", "INOCULAÇÃO BETONEIRA"),
        ("lona", "INOCULAÇÃO LONA"),
        ("sulco", "INOCULAÇÃO SULCO"),
    ]:
        escreve_linha(f"N° Questionários {titulo}", {k: v[cat] for k, v in r_forma.items()})
        escreve_linha(f"Percentual Questionários COM USO {titulo}",
                      {k: (v[cat] / n_total_geral) if n_total_geral else None for k, v in r_forma.items()})

    # Co e Mo: forma (semente/foliar)
    r_comomo = para_cada_grupo(lambda regs: {
        "semente": sum(1 for r in regs if "semente" in _norm(r.get("CoMo_Forma"))),
        "foliar": sum(1 for r in regs if "foliar" in _norm(r.get("CoMo_Forma"))),
    })
    pula_linha()
    escreve_linha("N° Questionários Co Mo SEMENTE", {k: v["semente"] for k, v in r_comomo.items()})
    escreve_linha("Percentual Questionários COM USO Co Mo SEMENTE",
                  {k: (v["semente"] / n_total_geral) if n_total_geral else None for k, v in r_comomo.items()})
    escreve_linha("N° Questionários Co Mo FOLIAR", {k: v["foliar"] for k, v in r_comomo.items()})
    escreve_linha("Percentual Questionários COM USO Co Mo FOLIAR",
                  {k: (v["foliar"] / n_total_geral) if n_total_geral else None for k, v in r_comomo.items()})


def _bloco_simples_texto(registros, campo, contem):
    """COM = registros cujo campo contem o texto 'contem' (case/acento
    insensiveis); usado p.ex. para 'classe do produto contem Inseticida'."""
    com = sum(1 for r in registros if contem.lower() in _norm(r.get(campo)))
    total = len(registros)
    sem = total - com
    return {
        "com": com, "sem": sem,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
    }


def _bloco_simples_preenchido(registros, campo):
    """COM = registros em que o campo (texto do alvo) esta preenchido."""
    com = sum(1 for r in registros if r.get(campo))
    total = len(registros)
    sem = total - com
    return {
        "com": com, "sem": sem,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
    }


# ---------------------------------------------------------------------------
# Funcao principal
# ---------------------------------------------------------------------------
def gerar_excel(registros: list[dict], filepath: str = "MesoIDR_Export.xlsx") -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _build_BD(wb)
    _build_total_pr(wb, registros)
    _build_medias_geral(wb, registros)

    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Teste local
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import random
    random.seed(7)

    MESOS = REGIOES_IDR
    MUNICS = {
        "Noroeste": ["Campo Mourao", "Umuarama", "Cianorte"],
        "Norte": ["Londrina", "Maringa", "Cornelio Procopio"],
        "Oeste": ["Cascavel", "Toledo", "Foz do Iguacu"],
        "Sudoeste": ["Pato Branco", "Francisco Beltrao"],
        "Centro Sul": ["Guarapuava", "Irati"],
        "Centro": ["Ponta Grossa", "Castro"],
        "Metropolitana e Litoral": ["Curitiba", "Paranagua"],
    }
    CULTIVARES = [" 50I52 RSF IPRO", " 5400 IPRO", " 5644 IPRO", " 6039 IPRO"]

    def _fake_reg(i):
        meso = random.choice(MESOS)
        munic = random.choice(MUNICS[meso])
        reg: dict[str, Any] = {
            "N": i, "Numero_Produtor": f"{i:04d}",
            "Meso_IDR": meso, "Regiao": meso, "Municipio": munic,
            "Area_Soja": round(random.uniform(50, 900), 1),
            "Cultivar": random.choice(CULTIVARES),
            "Bt": random.choice(["SIM", "NAO"]),
            "Produtividade": round(random.uniform(40, 85), 1),
            "Dt_Plantio": f"2024-10-{random.randint(1,28):02d}",
            "Adversidade": random.choice([None, "Seca", "Granizo"]),
            "Sinistro": random.choice(["SIM", "NAO"]),
            "Conhec_MID": random.choice(["SIM", "NAO"]),
            "Utiliza_MID": random.choice(["SIM", "NAO"]),
            "Conhec_MIP": random.choice(["SIM", "NAO"]),
            "Utiliza_MIP": random.choice(["SIM", "NAO"]),
            "Herb_Cl1": "Herbicida", "Herb_Alv1": "Folhas largas", "Herb_Nap1": 1,
            "Dess_Sim": "SIM", "Dess_Dt": "2024-09-20",
            "Dess_Cl": "Herbicida", "Dess_Alv1": "Folhas largas",
            "Tto_Semente": random.choice(["SIM", "NAO"]),
            "SAL_CB": random.choice(["SIM", "NAO"]),
            "Ctrl_Biol": random.choice(["SIM", "NAO"]),
            "Inoc_Usa": random.choice(["SIM", "NAO"]),
            "Inoc_Forma": "Via semente",
            "Coinoc": random.choice(["SIM", "NAO"]),
            "CoMo_Usa": random.choice(["SIM", "NAO"]),
            "CoMo_Forma": random.choice(["Via semente", "Foliar", None]),
        }
        for n in range(1, N_PULV + 1):
            reg[f"P{n}_DAE"] = random.randint(15, 90)
            reg[f"P{n}_Data"] = f"2024-{random.randint(10,12):02d}-{random.randint(1,28):02d}"
            reg[f"P{n}_Cl1"] = "Inseticida"
            reg[f"P{n}_Alv1"] = random.choice(PRAGAS)
        return reg

    registros = [_fake_reg(i) for i in range(1, 21)]
    out = gerar_excel(registros, "/home/claude/out/MesoIDR_Export_teste.xlsx")
    print(f"Gerado: {out} ({len(registros)} registros)")
